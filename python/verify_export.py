from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from common import ensure_dir, get_config, log, run_id


def newest_file(dir_path: Path) -> Path:
    entries = [entry for entry in dir_path.iterdir() if entry.is_file()]
    if not entries:
        raise RuntimeError(f"No files found in {dir_path}")
    entries.sort(key=lambda item: item.stat().st_mtime, reverse=True)
    return entries[0]


def parse_csv_file(file_path: Path) -> list[dict[str, Any]]:
    with file_path.open("r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f)
        return [dict(row) for row in reader]


def parse_xlsx_file(file_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers_row = next(rows_iter)
    except StopIteration:
        return []

    headers = [str(col) if col is not None else "" for col in headers_row]
    rows: list[dict[str, Any]] = []
    for row in rows_iter:
        item: dict[str, Any] = {}
        for idx, value in enumerate(row):
            key = headers[idx] if idx < len(headers) else f"col_{idx}"
            item[key] = value if value is not None else ""
        rows.append(item)
    return rows


def parse_file(file_path: Path) -> list[dict[str, Any]]:
    ext = file_path.suffix.lower()
    if ext in {".csv", ".txt"}:
        return parse_csv_file(file_path)
    if ext in {".xlsx", ".xls"}:
        return parse_xlsx_file(file_path)
    raise RuntimeError(f"Unsupported export file extension: {ext}")


def assert_required_columns(rows: list[dict[str, Any]], required_columns: list[str]) -> None:
    if not rows:
        raise RuntimeError("Export file has zero rows.")
    columns = list(rows[0].keys())
    missing = [required for required in required_columns if required not in columns]
    if missing:
        raise RuntimeError(f"Missing required columns: {', '.join(missing)}")


def main() -> None:
    job_id = run_id()
    config = get_config()
    ensure_dir(config.output_dir)

    if not config.downloads_dir.exists():
        raise RuntimeError(f"Downloads directory not found: {config.downloads_dir}")

    latest = newest_file(config.downloads_dir)
    rows = parse_file(latest)
    assert_required_columns(rows, config.required_columns)

    preview = rows[:3]
    normalized_path = config.output_dir / "latest.json"
    normalized_path.write_text(json.dumps(rows, indent=2, default=str), encoding="utf-8")

    log(
        job_id,
        "export verified",
        {
            "file": str(latest),
            "rowCount": len(rows),
            "preview": preview,
        },
    )
    log(job_id, "normalized json written", {"path": str(normalized_path)})


if __name__ == "__main__":
    main()
